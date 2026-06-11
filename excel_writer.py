"""
excel_writer.py — Gera/atualiza planilha Excel de controle de
pagamentos da Arezzo a partir dos dados extraidos do PDF.

Funcoes publicas:
- gerar_xlsx_bytes(...) -> bytes da planilha final

Layout da aba (14 colunas, sem a antiga coluna O "LIQ HOLERITE"):
  A: Nome
  B: MOTIVACIONAL          (input azul)
  C: HORA EXTRA            (input azul)
  D: DOMINGO               (input azul)
  E: HORA EXTRA TOTAL      (formula =C+D)
  F: COMISSAO              (formula inversa, comissionadas)
  G: SALARIO               (formula inversa, mensalistas)
  H: VALES                 (input azul)
  I: UNIODONTO             (input azul)
  J: PLANO DE SAUDE        (input azul)
  K: EMPRESTIMO            (input azul)
  L: VALE TRANSPORTE       (input azul)
  M: TOTAL DESCONTOS       (formula =H+I+J+K+L)
  N: LIQUIDO               (input azul, fundo verde, bold)

Formulas inversas:
  F = N - B - E + M   (so para comissionadas)
  G = N - B - E + M   (so para mensalistas)
"""

from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    CEF, CAIXA_LOJA, COMISSIONADAS, INCLUIR_CEF,
    nome_aba_por_data,
    COR_AZUL_INPUT, COR_PRETO_FORMULA, COR_VERDE_LIQUIDO,
    COR_AMARELO_HEADER, COR_CINZA_ESCURO, COR_BRANCO,
    LARGURAS, ALTURA_DADOS, ALTURA_TOTAL,
    ALTURA_HEADER_SECAO, ALTURA_HEADER_COLS,
    FORMATO_REAL,
)


# Layout das colunas
COLS = ["A", "B", "C", "D", "E", "F", "G",
        "H", "I", "J", "K", "L", "M", "N"]

HEADERS_COLUNAS = {
    "B": "MOTIVACIONAL",
    "C": "HORA EXTRA",
    "D": "DOMINGO",
    "E": "HORA EXTRA\nTOTAL",
    "F": "COMISSÃO",
    "G": "SALÁRIO",
    "H": "VALES",
    "I": "UNIODONTO",
    "J": "PLANO DE\nSAÚDE",
    "K": "EMPRÉSTIMO",
    "L": "VALE\nTRANSPORTE",
    "M": "TOTAL\nDESCONTOS",
    "N": "LÍQUIDO",
}

# Bordas finas pretas em todas as celulas
_BORDA = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)


# ============================================================
# Helpers de classificacao
# ============================================================

def classificar_grupo(nome, overrides=None):
    """Decide se a funcionaria vai pra CEF ou CAIXA DA LOJA."""
    if overrides and nome in overrides and "grupo" in overrides[nome]:
        return overrides[nome]["grupo"]
    if nome in INCLUIR_CEF:
        return "CEF"
    if nome in CEF:
        return "CEF"
    if nome in CAIXA_LOJA:
        return "CDL"
    # Desconhecida -> CDL por padrao (UI permite mover)
    return "CDL"


def eh_comissionada(nome, overrides=None):
    if overrides and nome in overrides and "comissionada" in overrides[nome]:
        return overrides[nome]["comissionada"]
    return nome in COMISSIONADAS


# ============================================================
# Estilos
# ============================================================

def _fonte_input():
    return Font(name="Arial", size=8, color=COR_AZUL_INPUT)


def _fonte_formula():
    return Font(name="Arial", size=8, color=COR_PRETO_FORMULA)


def _fonte_liquido():
    return Font(name="Arial", size=8, color=COR_AZUL_INPUT, bold=True)


def _fonte_nome():
    return Font(name="Arial", size=8, color=COR_PRETO_FORMULA)


def _fonte_total():
    return Font(name="Arial", size=9, color=COR_PRETO_FORMULA, bold=True)


def _fonte_header_secao():
    return Font(name="Arial", size=12, color=COR_PRETO_FORMULA, bold=True)


def _fonte_header_col():
    return Font(name="Arial", size=8, color=COR_BRANCO, bold=True)


def _aplicar_borda_e_centro(cel, centro=True):
    cel.border = _BORDA
    if centro:
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============================================================
# Escrita das linhas
# ============================================================

def _aplicar_larguras_e_layout(ws):
    """Define largura de coluna e congela painel."""
    for letra, largura in LARGURAS.items():
        if letra in COLS:
            ws.column_dimensions[letra].width = largura
    ws.sheet_view.showGridLines = False


def _escrever_cabecalho_secao(ws, linha, titulo, data_pag):
    """Linha amarela: 'CAIXA ECONOMICA' / 'CAIXA DA LOJA (DINHEIRO)' + data."""
    fill = PatternFill("solid", fgColor=COR_AMARELO_HEADER)
    for col in COLS:
        cel = ws[f"{col}{linha}"]
        cel.fill = fill
        cel.border = _BORDA
        cel.font = _fonte_header_secao()
    cel_a = ws[f"A{linha}"]
    cel_a.value = titulo
    cel_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cel_b = ws[f"B{linha}"]
    cel_b.value = "Data pagamento:"
    cel_b.font = Font(name="Arial", size=10, bold=True)
    cel_b.alignment = Alignment(horizontal="right", vertical="center")
    cel_c = ws[f"C{linha}"]
    cel_c.value = data_pag.strftime("%d/%m/%Y")
    cel_c.font = Font(name="Arial", size=10, bold=True)
    cel_c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = ALTURA_HEADER_SECAO


def _escrever_cabecalho_colunas(ws, linha):
    fill = PatternFill("solid", fgColor=COR_CINZA_ESCURO)
    for col in COLS:
        cel = ws[f"{col}{linha}"]
        cel.fill = fill
        cel.border = _BORDA
        cel.font = _fonte_header_col()
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, texto in HEADERS_COLUNAS.items():
        ws[f"{col}{linha}"].value = texto
    ws.row_dimensions[linha].height = ALTURA_HEADER_COLS


def _escrever_linha_funcionaria(ws, linha, nome, dados, comissionada):
    """
    Preenche uma linha de dados.
    - dados: dict do parser (motivacional, he, domingo, vales, uniodonto,
      plano_saude, emprestimo, vale_transporte, liquido)
    - comissionada: True -> formula vai em F, False -> formula vai em G
    """
    # A: Nome
    cel_a = ws[f"A{linha}"]
    cel_a.value = nome
    cel_a.font = _fonte_nome()
    cel_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cel_a.border = _BORDA

    # Inputs (azul). 0,00 vira "" pra ficar visualmente mais limpo? Nao.
    # A spec diz: formatar TODAS celulas inclusive vazias. Manter 0,00 explicitos
    # so quando ha valor; senao deixar value=None mas com format.
    inputs = {
        "B": dados.get("motivacional", 0),
        "C": dados.get("he", 0),
        "D": dados.get("domingo", 0),
        "H": dados.get("vales", 0),
        "I": dados.get("uniodonto", 0),
        "J": dados.get("plano_saude", 0),
        "K": dados.get("emprestimo", 0),
        "L": dados.get("vale_transporte", 0),
    }
    for col, valor in inputs.items():
        cel = ws[f"{col}{linha}"]
        cel.value = valor if valor else None
        cel.font = _fonte_input()
        cel.number_format = FORMATO_REAL
        _aplicar_borda_e_centro(cel)

    # E: HORA EXTRA TOTAL (formula)
    cel_e = ws[f"E{linha}"]
    cel_e.value = f"=IFERROR(C{linha}+D{linha},0)"
    cel_e.font = _fonte_formula()
    cel_e.number_format = FORMATO_REAL
    _aplicar_borda_e_centro(cel_e)

    # F ou G: COMISSAO/SALARIO (formula inversa = N - B - E + M)
    formula_inv = f"=IFERROR(N{linha}-B{linha}-E{linha}+M{linha},0)"
    for col in ("F", "G"):
        cel = ws[f"{col}{linha}"]
        if (col == "F" and comissionada) or (col == "G" and not comissionada):
            cel.value = formula_inv
        else:
            cel.value = None
        cel.font = _fonte_formula()
        cel.number_format = FORMATO_REAL
        _aplicar_borda_e_centro(cel)

    # M: TOTAL DESCONTOS (formula)
    cel_m = ws[f"M{linha}"]
    cel_m.value = f"=IFERROR(H{linha}+I{linha}+J{linha}+K{linha}+L{linha},0)"
    cel_m.font = _fonte_formula()
    cel_m.number_format = FORMATO_REAL
    _aplicar_borda_e_centro(cel_m)

    # N: LIQUIDO (input do PDF, fundo verde, bold, azul)
    cel_n = ws[f"N{linha}"]
    cel_n.value = dados.get("liquido", 0)
    cel_n.font = _fonte_liquido()
    cel_n.number_format = FORMATO_REAL
    cel_n.fill = PatternFill("solid", fgColor=COR_VERDE_LIQUIDO)
    _aplicar_borda_e_centro(cel_n)

    ws.row_dimensions[linha].height = ALTURA_DADOS


def _escrever_linha_total(ws, linha, linha_inicio, linha_fim):
    """Linha TOTAL amarela com SUM() de cada coluna."""
    fill = PatternFill("solid", fgColor=COR_AMARELO_HEADER)
    cel_a = ws[f"A{linha}"]
    cel_a.value = "TOTAL"
    cel_a.font = _fonte_total()
    cel_a.fill = fill
    cel_a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cel_a.border = _BORDA
    for col in COLS[1:]:
        cel = ws[f"{col}{linha}"]
        cel.value = f"=SUM({col}{linha_inicio}:{col}{linha_fim})"
        cel.font = _fonte_total()
        cel.fill = fill
        cel.number_format = FORMATO_REAL
        _aplicar_borda_e_centro(cel)
    # Coluna N (Liquido) na linha TOTAL fica com fundo verde, nao amarelo
    cel_n = ws[f"N{linha}"]
    cel_n.fill = PatternFill("solid", fgColor=COR_VERDE_LIQUIDO)
    ws.row_dimensions[linha].height = ALTURA_TOTAL


# ============================================================
# Montagem da aba completa
# ============================================================

def _ordenar_alfabetico(nomes):
    import unicodedata
    def chave(s):
        return "".join(
            c for c in unicodedata.normalize("NFKD", s)
            if not unicodedata.combining(c)
        ).lower()
    return sorted(nomes, key=chave)


def _separar_grupos(dados, overrides=None):
    """Devolve (lista_cef, lista_cdl) com nomes ordenados alfabeticamente."""
    cef, cdl = [], []
    for nome in dados:
        grupo = classificar_grupo(nome, overrides)
        if grupo == "CEF":
            cef.append(nome)
        else:
            cdl.append(nome)
    return _ordenar_alfabetico(cef), _ordenar_alfabetico(cdl)


def _limpar_aba(ws):
    """Apaga conteudo da aba mantendo a referencia do worksheet."""
    ws.delete_rows(1, ws.max_row + 1 if ws.max_row else 1)


def _escrever_aba(ws, data_pag, dados, overrides=None):
    """
    Monta a aba inteira:
      Linha 2: header secao CEF
      Linha 3: headers de coluna
      Linha 4..N: funcionarias CEF (alfabetico)
      Linha N+1: TOTAL CEF
      Linha N+4: header secao CDL
      Linha N+5: headers de coluna
      Linha N+6..M: funcionarias CDL (alfabetico)
      Linha M+1: TOTAL CDL
    """
    _aplicar_larguras_e_layout(ws)
    cef, cdl = _separar_grupos(dados, overrides)

    # Linha 1 fica em branco (margem visual)
    ws.row_dimensions[1].height = 7

    # ====== Secao CEF ======
    linha_header_secao_cef = 2
    linha_header_cols_cef = 3
    linha_inicio_cef = 4
    _escrever_cabecalho_secao(ws, linha_header_secao_cef, "CAIXA ECONÔMICA", data_pag)
    _escrever_cabecalho_colunas(ws, linha_header_cols_cef)

    for i, nome in enumerate(cef):
        linha = linha_inicio_cef + i
        comiss = eh_comissionada(nome, overrides)
        _escrever_linha_funcionaria(ws, linha, nome, dados[nome], comiss)
    linha_total_cef = linha_inicio_cef + len(cef)
    if cef:
        _escrever_linha_total(ws, linha_total_cef, linha_inicio_cef, linha_total_cef - 1)

    # ====== Linhas em branco separador ======
    linha_proxima = linha_total_cef + 2 if cef else linha_inicio_cef + 1

    # ====== Secao CDL ======
    linha_header_secao_cdl = linha_proxima + 1
    linha_header_cols_cdl = linha_header_secao_cdl + 1
    linha_inicio_cdl = linha_header_cols_cdl + 1
    _escrever_cabecalho_secao(ws, linha_header_secao_cdl, "CAIXA DA LOJA (DINHEIRO)", data_pag)
    _escrever_cabecalho_colunas(ws, linha_header_cols_cdl)

    for i, nome in enumerate(cdl):
        linha = linha_inicio_cdl + i
        comiss = eh_comissionada(nome, overrides)
        _escrever_linha_funcionaria(ws, linha, nome, dados[nome], comiss)
    linha_total_cdl = linha_inicio_cdl + len(cdl)
    if cdl:
        _escrever_linha_total(ws, linha_total_cdl, linha_inicio_cdl, linha_total_cdl - 1)


# ============================================================
# Ordem das abas (cronologica)
# ============================================================

def _indice_mes_ano(nome_aba):
    """'Junho 2026' -> (2026, 6). Devolve None se nao reconhecer."""
    meses = {
        "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4,
        "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,
        "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12,
    }
    partes = nome_aba.strip().split()
    if len(partes) != 2:
        return None
    mes_nome, ano_str = partes
    mes = meses.get(mes_nome)
    if not mes:
        return None
    try:
        ano = int(ano_str)
    except ValueError:
        return None
    return (ano, mes)


def _ordenar_abas_cronologico(wb):
    """Reordena abas em ordem cronologica. Abas com nome nao-padrao vao no fim."""
    nomes = list(wb.sheetnames)
    cronologicas = [(n, _indice_mes_ano(n)) for n in nomes]
    cronologicas.sort(key=lambda x: (x[1] is None, x[1] or (9999, 99), x[0]))
    nova_ordem = [n for n, _ in cronologicas]
    wb._sheets = [wb[n] for n in nova_ordem]


# ============================================================
# API publica
# ============================================================

def gerar_xlsx_de_holerites(linhas, data_pag, planilha_existente_bytes=None,
                            tipo="regular", competencia_label=None):
    """
    Gera/atualiza planilha a partir de uma lista de holerites do banco.

    `linhas`: list de dicts com:
      nome, banco_nome, banco_ordem, comissionada, motivacional, he, domingo,
      vales, uniodonto, plano_saude, emprestimo, vale_transporte, liquido

    `tipo`: 'regular' | '13_1' | '13_2'
    `competencia_label`: texto curto pra mostrar no header da secao
                        (ex: "Maio 2026", "Ano 2026", "13º 1ª parcela")

    Cada banco vira uma secao da aba (multi-banco).
    """
    if planilha_existente_bytes:
        wb = openpyxl.load_workbook(BytesIO(planilha_existente_bytes))
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    nome_aba = nome_aba_por_data(data_pag)
    if tipo == "13_1":
        nome_aba = f"{nome_aba} - 13o 1a"
    elif tipo == "13_2":
        nome_aba = f"{nome_aba} - 13o 2a"
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        _limpar_aba(ws)
    else:
        ws = wb.create_sheet(title=nome_aba)

    _aplicar_larguras_e_layout(ws)
    ws.row_dimensions[1].height = 7

    # Agrupa por banco (ordenado por banco_ordem)
    import unicodedata
    def chave(s):
        return "".join(c for c in unicodedata.normalize("NFKD", s)
                       if not unicodedata.combining(c)).lower()
    grupos = {}
    for l in linhas:
        grupos.setdefault((l["banco_ordem"], l["banco_nome"]), []).append(l)

    linha = 2
    for (ordem, banco_nome) in sorted(grupos.keys()):
        grupo = grupos[(ordem, banco_nome)]
        if not grupo:
            continue
        grupo_ord = sorted(grupo, key=lambda l: chave(l["nome"]))

        titulo_secao = banco_nome.upper()
        if competencia_label:
            titulo_secao += f"   |   {competencia_label}"
        _escrever_cabecalho_secao(ws, linha, titulo_secao, data_pag)
        linha += 1
        _escrever_cabecalho_colunas(ws, linha)
        linha += 1
        linha_inicio = linha
        for l in grupo_ord:
            dados_celula = {
                "motivacional": l["motivacional"],
                "he": l["he"],
                "domingo": l["domingo"],
                "vales": l["vales"],
                "uniodonto": l["uniodonto"],
                "plano_saude": l["plano_saude"],
                "emprestimo": l["emprestimo"],
                "vale_transporte": l["vale_transporte"],
                "liquido": l["liquido"],
            }
            _escrever_linha_funcionaria(ws, linha, l["nome"], dados_celula, l["comissionada"])
            linha += 1
        _escrever_linha_total(ws, linha, linha_inicio, linha - 1)
        linha += 3  # separador entre secoes

    _ordenar_abas_cronologico(wb)
    wb.active = wb.sheetnames.index(nome_aba)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_xlsx_bytes(dados, data_pag, overrides=None, planilha_existente_bytes=None):
    """
    Gera/atualiza a planilha e devolve os bytes do .xlsx final.

    Parametros
    ----------
    dados : dict { nome -> dict }
        Saida do parser.extrair_holerites.
    data_pag : datetime.date
        Data de pagamento (ex: 05/06/2026 -> aba "Junho 2026").
    overrides : dict opcional
        Ajustes do usuario na UI:
          { nome: {"grupo": "CEF"|"CDL", "comissionada": bool} }
    planilha_existente_bytes : bytes opcional
        Bytes da planilha atual; se passado, abre e adiciona/substitui
        a aba do mes. Se nao, cria do zero.

    Retorno
    -------
    bytes da planilha .xlsx
    """
    if planilha_existente_bytes:
        wb = openpyxl.load_workbook(BytesIO(planilha_existente_bytes))
    else:
        wb = Workbook()
        # Remove a aba padrao "Sheet"
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    nome_aba = nome_aba_por_data(data_pag)
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        _limpar_aba(ws)
    else:
        ws = wb.create_sheet(title=nome_aba)

    _escrever_aba(ws, data_pag, dados, overrides)
    _ordenar_abas_cronologico(wb)
    wb.active = wb.sheetnames.index(nome_aba)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
